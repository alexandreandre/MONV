#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════╗
║  GENERATEUR DE DATASET PME - PROSPECTION COMMERCIALE           ║
║  Source : API Recherche d'Entreprises (data.gouv.fr)           ║
║  Gratuit, sans clé API, données publiques INSEE                ║
╚══════════════════════════════════════════════════════════════════╝

Usage:
    pip install requests pandas openpyxl
    python prospection_pme.py

Le script génère un fichier Excel "prospection_pme.xlsx" avec toutes
les PME françaises correspondant à tes critères.
"""

import requests
import pandas as pd
import time
import sys
from datetime import datetime

# ============================================================
# CONFIGURATION - Modifie ces paramètres selon tes besoins
# ============================================================

CONFIG = {
    # Tranches d'effectifs salariés (codes INSEE)
    # "12" = 20-49, "21" = 50-99, "22" = 100-199,
    # "31" = 200-249, "32" = 250-499
    "tranches_effectifs": ["12", "21", "22", "31", "32"],

    # Codes NAF (secteurs d'activité) - les plus pertinents pour ton offre
    # Tu peux en ajouter/retirer. Liste complète: https://www.insee.fr/fr/information/2120875
    "secteurs": {
        # --- INDUSTRIE / MANUFACTURING ---
        "10": "Industrie alimentaire",
        "20": "Industrie chimique",
        "25": "Fabrication produits métalliques",
        "26": "Fabrication produits informatiques/électroniques",
        "27": "Fabrication équipements électriques",
        "28": "Fabrication machines/équipements",
        "29": "Industrie automobile",

        # --- SERVICES B2B / CONSEIL ---
        "62": "Programmation/conseil informatique",
        "69": "Activités juridiques et comptables",
        "70": "Activités des sièges sociaux/conseil gestion",
        "71": "Architecture/ingénierie/contrôle technique",
        "73": "Publicité et études de marché",
        "74": "Autres activités spécialisées scientifiques",
        "78": "Activités liées à l'emploi",
        "80": "Enquêtes et sécurité",
        "82": "Services administratifs de bureau",

        # --- TECH / SAAS / DIGITAL ---
        "58": "Edition (logiciels inclus)",
        "61": "Télécommunications",
        "63": "Services d'information",

        # --- SANTÉ / PHARMA ---
        "21": "Industrie pharmaceutique",
        "32": "Autres industries manufacturières",
        "86": "Activités pour la santé humaine",

        # --- AUTRES SECTEURS À FORT BESOIN DOCUMENTAIRE ---
        "41": "Construction de bâtiments",
        "42": "Génie civil",
        "43": "Travaux de construction spécialisés",
        "64": "Services financiers (hors assurance)",
        "65": "Assurance",
        "66": "Activités auxiliaires finance/assurance",
        "85": "Enseignement",
    },

    # Nombre max de résultats par requête (l'API plafonne à 10000)
    "max_par_requete": 25,

    # Nombre de pages à parcourir par secteur (25 résultats/page)
    # Augmente pour plus de résultats (attention au rate limit: 7 req/s)
    "pages_par_secteur": 40,

    # Délai entre requêtes (secondes) pour respecter le rate limit
    "delai_requetes": 0.2,

    # Fichier de sortie
    "fichier_sortie": "prospection_pme.xlsx",
}

# ============================================================
# API CONFIGURATION
# ============================================================

BASE_URL = "https://recherche-entreprises.api.gouv.fr/search"

TRANCHE_LABELS = {
    "00": "0 salarié",
    "01": "1-2 salariés",
    "02": "3-5 salariés",
    "03": "6-9 salariés",
    "11": "10-19 salariés",
    "12": "20-49 salariés",
    "21": "50-99 salariés",
    "22": "100-199 salariés",
    "31": "200-249 salariés",
    "32": "250-499 salariés",
    "41": "500-999 salariés",
    "42": "1000-1999 salariés",
    "51": "2000-4999 salariés",
    "52": "5000-9999 salariés",
    "53": "10000+ salariés",
}


def fetch_entreprises(section_naf, page=1):
    """Interroge l'API pour un code NAF (section = 2 premiers chiffres)."""
    params = {
        "section_activite_principale": section_naf,
        "tranche_effectif_salarie": ",".join(CONFIG["tranches_effectifs"]),
        "etat_administratif": "A",  # Entreprises actives uniquement
        "page": page,
        "per_page": CONFIG["max_par_requete"],
    }

    try:
        resp = requests.get(BASE_URL, params=params, timeout=15)
        if resp.status_code == 429:
            print("  ⏳ Rate limit atteint, pause de 5s...")
            time.sleep(5)
            return fetch_entreprises(section_naf, page)
        resp.raise_for_status()
        return resp.json()
    except requests.RequestException as e:
        print(f"  ❌ Erreur API: {e}")
        return None


def extract_data(result):
    """Extrait les champs utiles d'un résultat API."""
    siege = result.get("siege", {})
    dirigeants = result.get("dirigeants", [])

    # Trouver le dirigeant principal (président, DG, gérant)
    dirigeant_nom = ""
    dirigeant_fonction = ""
    for d in dirigeants:
        qualite = (d.get("qualite") or d.get("fonction") or "").lower()
        if any(k in qualite for k in ["président", "directeur général",
                                        "gérant", "pdg", "dirigeant"]):
            prenom = d.get("prenom") or d.get("prenoms") or ""
            nom = d.get("nom") or ""
            dirigeant_nom = f"{prenom} {nom}".strip()
            dirigeant_fonction = d.get("qualite") or d.get("fonction") or ""
            break
    if not dirigeant_nom and dirigeants:
        d = dirigeants[0]
        prenom = d.get("prenom") or d.get("prenoms") or ""
        nom = d.get("nom") or ""
        dirigeant_nom = f"{prenom} {nom}".strip()
        dirigeant_fonction = d.get("qualite") or d.get("fonction") or ""

    tranche = result.get("tranche_effectif_salarie") or ""
    tranche_label = TRANCHE_LABELS.get(tranche, tranche)

    # Construire l'adresse complète
    adresse_parts = []
    for field in ["numero_voie", "type_voie", "libelle_voie"]:
        val = siege.get(field)
        if val:
            adresse_parts.append(str(val))
    adresse_ligne1 = " ".join(adresse_parts)

    return {
        "nom_entreprise": result.get("nom_complet", ""),
        "siren": result.get("siren", ""),
        "siret_siege": siege.get("siret", ""),
        "activite_principale": result.get("activite_principale", ""),
        "libelle_activite": result.get("libelle_activite_principale", ""),
        "tranche_effectif_code": tranche,
        "tranche_effectif": tranche_label,
        "nature_juridique": result.get("nature_juridique", ""),
        "date_creation": result.get("date_creation", ""),
        "dirigeant_nom": dirigeant_nom,
        "dirigeant_fonction": dirigeant_fonction,
        "adresse": adresse_ligne1,
        "code_postal": siege.get("code_postal", ""),
        "ville": siege.get("libelle_commune", ""),
        "region": siege.get("region", "") if siege.get("region") else "",
        "departement": siege.get("departement", "") if siege.get("departement") else "",
        "lien_annuaire": f"https://annuaire-entreprises.data.gouv.fr/entreprise/{result.get('siren', '')}",
    }


def main():
    print("=" * 65)
    print("  🏢 GÉNÉRATEUR DE DATASET PME - PROSPECTION COMMERCIALE")
    print("=" * 65)
    print(f"  Tranches: {', '.join(TRANCHE_LABELS[t] for t in CONFIG['tranches_effectifs'])}")
    print(f"  Secteurs: {len(CONFIG['secteurs'])} secteurs ciblés")
    print(f"  Pages/secteur: {CONFIG['pages_par_secteur']}")
    print()

    all_results = []
    seen_sirens = set()
    total_sections = len(CONFIG["secteurs"])

    for idx, (code_naf, label) in enumerate(CONFIG["secteurs"].items(), 1):
        # L'API utilise les "sections" NAF (lettre), mais on peut aussi
        # filtrer par les 2 premiers chiffres du code NAF via section_activite_principale
        # Mapping: codes numériques vers sections lettres
        naf_to_section = {
            "10": "C", "20": "C", "21": "C", "25": "C", "26": "C",
            "27": "C", "28": "C", "29": "C", "32": "C",
            "41": "F", "42": "F", "43": "F",
            "58": "J", "61": "J", "62": "J", "63": "J",
            "64": "K", "65": "K", "66": "K",
            "69": "M", "70": "M", "71": "M", "73": "M", "74": "M",
            "78": "N", "80": "N", "82": "N",
            "85": "P",
            "86": "Q",
        }
        section = naf_to_section.get(code_naf)
        if not section:
            continue

        print(f"[{idx}/{total_sections}] 📂 Section {section} - {label}...")

        for page in range(1, CONFIG["pages_par_secteur"] + 1):
            data = fetch_entreprises(section, page)
            if not data:
                break

            results = data.get("results", [])
            total_api = data.get("total_results", 0)

            if not results:
                break

            for r in results:
                siren = r.get("siren")
                if siren and siren not in seen_sirens:
                    # Filtrer par code NAF exact (2 premiers chiffres)
                    activite = r.get("activite_principale", "")
                    if activite and activite[:2] == code_naf:
                        seen_sirens.add(siren)
                        all_results.append(extract_data(r))

            time.sleep(CONFIG["delai_requetes"])

            # S'arrêter si on a parcouru tous les résultats
            if page * CONFIG["max_par_requete"] >= min(total_api, 10000):
                break

        print(f"  ✅ {len([r for r in all_results if r['activite_principale'][:2] == code_naf])} entreprises trouvées (total cumulé: {len(all_results)})")

        # Dédupliquer les sections déjà vues pour éviter les requêtes redondantes
        # (plusieurs codes NAF peuvent être dans la même section)

    print()
    print(f"📊 Total brut: {len(all_results)} entreprises uniques")

    if not all_results:
        print("❌ Aucun résultat. Vérifie ta connexion internet.")
        sys.exit(1)

    # Créer le DataFrame
    df = pd.DataFrame(all_results)

    # Trier par taille décroissante puis par nom
    tranche_order = {"32": 0, "31": 1, "22": 2, "21": 3, "12": 4}
    df["_sort"] = df["tranche_effectif_code"].map(tranche_order).fillna(5)
    df = df.sort_values(["_sort", "nom_entreprise"]).drop(columns=["_sort"])

    # Sauvegarder en Excel
    output_file = CONFIG["fichier_sortie"]
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="PME_Prospection", index=False)

        # Ajuster la largeur des colonnes
        ws = writer.sheets["PME_Prospection"]
        col_widths = {
            "A": 35, "B": 12, "C": 16, "D": 10, "E": 35,
            "F": 8, "G": 18, "H": 12, "I": 12, "J": 25,
            "K": 20, "L": 30, "M": 8, "N": 18, "O": 20,
            "P": 15, "Q": 50,
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Mettre en forme l'en-tête
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Figer la première ligne
        ws.freeze_panes = "A2"

        # Ajouter un onglet de stats
        stats_data = df.groupby(["libelle_activite", "tranche_effectif"]).size().reset_index(name="nb_entreprises")
        stats_data.to_excel(writer, sheet_name="Stats_par_secteur", index=False)

    print(f"✅ Fichier généré: {output_file}")
    print(f"   → {len(df)} entreprises")
    print(f"   → {df['libelle_activite'].nunique()} secteurs d'activité")
    print(f"   → {df['ville'].nunique()} villes")
    print()

    # Résumé par tranche
    print("📈 Répartition par tranche d'effectif:")
    for tranche, count in df["tranche_effectif"].value_counts().items():
        print(f"   {tranche}: {count}")

    print()
    print("🎯 Prochaines étapes:")
    print("   1. Ouvre le fichier Excel et filtre par secteur/ville")
    print("   2. Cherche les dirigeants sur LinkedIn via le nom + entreprise")
    print("   3. Utilise Kaspr/Phantombuster pour enrichir les emails")
    print("   4. Lance ta séquence de prospection!")


if __name__ == "__main__":
    main()