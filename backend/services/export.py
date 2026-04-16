"""
Couche 4 — Sortie et livraison
Génère des fichiers Excel et CSV à partir des résultats de recherche.
"""

from __future__ import annotations

import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from config import settings
from models.schemas import SearchResults

COLUMN_LABELS = {
    "nom": "Dénomination / raison sociale",
    "siren": "SIREN (9 chiffres)",
    "siret": "SIRET (établissement siège)",
    "activite_principale": "Code NAF / APE",
    "libelle_activite": "Libellé d'activité (NAF)",
    "adresse": "Adresse (voie)",
    "code_postal": "Code postal",
    "ville": "Commune",
    "region": "Région",
    "departement": "Département",
    "tranche_effectif": "Effectif — code tranche INSEE",
    "effectif_label": "Effectif (tranche salariés)",
    "date_creation": "Date de création (INSEE)",
    "categorie_entreprise": "Catégorie d'entreprise (INSEE)",
    "forme_juridique": "Forme juridique (code / libellé)",
    "dirigeant_nom": "Dirigeant — nom",
    "dirigeant_prenom": "Dirigeant — prénom",
    "dirigeant_fonction": "Dirigeant — qualité / fonction",
    "annee_dernier_ca": "Exercice (année du dernier CA)",
    "date_cloture_exercice": "Date de clôture d'exercice",
    "chiffre_affaires": "Chiffre d'affaires (€)",
    "resultat_net": "Résultat net (€)",
    "marge_brute": "Marge brute (€)",
    "ebe": "Excédent brut d'exploitation — EBE (€)",
    "capitaux_propres": "Capitaux propres (€)",
    "effectif_financier": "Effectif (déclaration comptable)",
    "capital_social": "Capital social (€)",
    "email": "Email",
    "telephone": "Téléphone",
    "site_web": "Site web",
    "lien_annuaire": "Lien fiche entreprise (Annuaire des entreprises)",
    "google_maps_url": "Lien Google Maps",
}

INTENT_LABELS_FR: dict[str, str] = {
    "recherche_entreprise": "Liste / ciblage d'entreprises",
    "recherche_dirigeant": "Recherche orientée dirigeants",
    "enrichissement": "Enrichissement de données",
    "hors_scope": "Hors périmètre métier",
    "salutation": "Échange",
    "meta_question": "Question sur l'outil",
}

ENTITY_LABELS_FR: dict[str, str] = {
    "localisation": "Zone / localisation",
    "departement": "Département (critère)",
    "region": "Région (critère)",
    "secteur": "Secteur / thématique",
    "code_naf": "Code NAF visé",
    "taille_min": "Effectif minimum (approx.)",
    "taille_max": "Effectif maximum (approx.)",
    "ca_min": "CA minimum (€)",
    "ca_max": "CA maximum (€)",
    "date_creation_apres": "Création après le",
    "date_creation_avant": "Création avant le",
    "mots_cles": "Mots-clés",
    "forme_juridique": "Forme juridique (critère)",
}


def _format_entity_value(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, list):
        return ", ".join(str(x) for x in v if x is not None and str(x).strip())
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return f"{v:,.0f}".replace(",", " ")
    return str(v).strip()


def _entities_summary_lines(entities: dict[str, Any] | None) -> list[tuple[str, str]]:
    if not entities:
        return []
    lines: list[tuple[str, str]] = []
    for key, label in ENTITY_LABELS_FR.items():
        if key not in entities:
            continue
        val = _format_entity_value(entities.get(key))
        if not val:
            continue
        lines.append((label, val))
    return lines


def _results_to_dataframe(results: SearchResults, *, rename: bool = True) -> pd.DataFrame:
    """Convertit les résultats en DataFrame (colonnes internes, renommage optionnel)."""
    data = [r.model_dump() for r in results.results]
    df = pd.DataFrame(data)
    if df.empty:
        return df

    cols_to_keep = [c for c in results.columns if c in df.columns]
    if "lien_annuaire" in df.columns and "lien_annuaire" not in cols_to_keep:
        cols_to_keep.append("lien_annuaire")
    if "effectif_label" in df.columns and "effectif_label" not in cols_to_keep:
        cols_to_keep.append("effectif_label")

    df = df[cols_to_keep]
    if rename:
        df = df.rename(columns={c: COLUMN_LABELS.get(c, c) for c in df.columns})
    return df


def _style_header_row(ws, row_idx: int = 1) -> None:
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for cell in ws[row_idx]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _autosize_columns(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = col[0].column_letter
        maxlen = 0
        for cell in col:
            if cell.value is not None:
                maxlen = max(maxlen, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = min(maxlen + 2, max_width)


def _append_title(ws, row: int, title: str) -> int:
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(bold=True, size=11)
    return row + 1


def _append_dataframe_block(ws, start_row: int, df: pd.DataFrame, *, with_header: bool = True) -> int:
    if df.empty:
        ws.cell(row=start_row, column=1, value="—")
        return start_row + 1
    rows = list(dataframe_to_rows(df, index=False, header=with_header))
    for offset, row in enumerate(rows):
        ridx = start_row + offset
        for j, value in enumerate(row, start=1):
            ws.cell(row=ridx, column=j, value=value)
    return start_row + len(rows)


def _build_synthese_dataframes(
    df: pd.DataFrame,
    intent: str,
    entities: dict[str, Any] | None,
) -> list[tuple[str, pd.DataFrame]]:
    """Blocs (titre, tableau) pour l'onglet Synthèse."""
    blocks: list[tuple[str, pd.DataFrame]] = []
    ent = entities or {}

    crit_lines = _entities_summary_lines(ent)
    if crit_lines:
        crit = pd.DataFrame(crit_lines, columns=["Critère extrait", "Valeur"])
    else:
        crit = pd.DataFrame(
            {
                "Critère extrait": ["—"],
                "Valeur": ["Aucun critère structuré disponible pour cette recherche."],
            }
        )
    blocks.append(("Critères issus de ta requête", crit))

    if df.empty:
        blocks.append(("Analyse des résultats", pd.DataFrame({"Message": ["Aucune ligne à agréger."]})))
        return blocks

    n = len(df)
    dept_u = df["departement"].nunique(dropna=True) if "departement" in df.columns else 0
    ville_u = df["ville"].nunique(dropna=True) if "ville" in df.columns else 0
    naf_u = df["activite_principale"].nunique(dropna=True) if "activite_principale" in df.columns else 0

    vue = pd.DataFrame(
        {
            "Indicateur": [
                "Nombre d'établissements / lignes",
                "Départements distincts",
                "Communes distinctes",
                "Codes NAF distincts",
            ],
            "Valeur": [n, int(dept_u), int(ville_u), int(naf_u)],
        }
    )
    blocks.append(("Vue d'ensemble", vue))

    def _top_counts(col: str, label: str, top: int = 20) -> pd.DataFrame:
        if col not in df.columns:
            return pd.DataFrame()
        s = df[col].dropna().astype(str).str.strip()
        s = s[s != ""]
        if s.empty:
            return pd.DataFrame()
        vc = s.value_counts().head(top).reset_index()
        vc.columns = [label, "Nombre"]
        return vc

    t = _top_counts("departement", "Département")
    if not t.empty:
        blocks.append((f"Répartition par département (top {len(t)})", t))

    t = _top_counts("ville", "Commune")
    if not t.empty:
        blocks.append((f"Répartition par commune (top {len(t)})", t))

    t = _top_counts("libelle_activite", "Libellé d'activité")
    if t.empty and "activite_principale" in df.columns:
        t = _top_counts("activite_principale", "Code NAF")
    if not t.empty:
        blocks.append((f"Répartition par activité (top {len(t)})", t))

    eff_col = "effectif_label" if "effectif_label" in df.columns else "tranche_effectif"
    if eff_col in df.columns:
        s = df[eff_col].fillna("(non renseigné)").astype(str).str.strip()
        s = s.replace("", "(non renseigné)")
        vc = s.value_counts().reset_index()
        vc.columns = ["Tranche d'effectif", "Nombre"]
        blocks.append(("Effectifs (tranches INSEE)", vc))

    if intent == "recherche_dirigeant" and "dirigeant_nom" in df.columns:
        filled = df["dirigeant_nom"].notna() & (df["dirigeant_nom"].astype(str).str.strip() != "")
        qual = pd.DataFrame(
            {
                "Indicateur": [
                    "Lignes avec nom de dirigeant renseigné",
                    "Lignes sans dirigeant identifié",
                ],
                "Nombre": [int(filled.sum()), int((~filled).sum())],
            }
        )
        blocks.append(("Qualité — dirigeants", qual))
        if "dirigeant_fonction" in df.columns:
            ff = df.loc[filled, "dirigeant_fonction"].fillna("(non renseigné)").astype(str).str.strip()
            ff = ff.replace("", "(non renseigné)")
            if not ff.empty:
                fvc = ff.value_counts().head(15).reset_index()
                fvc.columns = ["Fonction / qualité", "Nombre"]
                blocks.append(("Dirigeants par fonction (échantillon)", fvc))

    if "chiffre_affaires" in df.columns:
        ca = pd.to_numeric(df["chiffre_affaires"], errors="coerce")
        nn = ca.notna().sum()
        if nn > 0:
            stats = pd.DataFrame(
                {
                    "Indicateur": [
                        "Lignes avec CA renseigné",
                        "CA minimum (€)",
                        "CA maximum (€)",
                        "CA médian (€)",
                    ],
                    "Valeur": [
                        int(nn),
                        float(ca.min()),
                        float(ca.max()),
                        float(ca.median()),
                    ],
                }
            )
            blocks.append(("Chiffre d'affaires (lignes enrichies)", stats))

    if "google_maps_url" in df.columns:
        gm = df["google_maps_url"].notna() & (df["google_maps_url"].astype(str).str.strip() != "")
        blocks.append(
            (
                "Complément local (Google Maps)",
                pd.DataFrame(
                    {
                        "Indicateur": ["Lignes avec lien Maps"],
                        "Nombre": [int(gm.sum())],
                    }
                ),
            )
        )

    if "email" in df.columns:
        em = df["email"].notna() & (df["email"].astype(str).str.strip() != "")
        if em.any():
            blocks.append(
                (
                    "Contacts",
                    pd.DataFrame(
                        {
                            "Indicateur": ["Lignes avec email renseigné"],
                            "Nombre": [int(em.sum())],
                        }
                    ),
                )
            )

    return blocks


def _fill_info_sheet(
    ws,
    *,
    query_text: str,
    intent: str,
    entities: dict[str, Any] | None,
    total: int,
    credits_used: int | None,
) -> None:
    bold = Font(bold=True)
    row = 1
    pairs: list[tuple[str, Any]] = [
        ("Requête d'origine", query_text or "—"),
        ("Date d'export", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Type d'intention", INTENT_LABELS_FR.get(intent, intent)),
        ("Nombre de lignes exportées", total),
    ]
    if credits_used is not None:
        pairs.append(("Crédits associés à cette recherche", credits_used))
    pairs.append(("Généré par", "MONV — monv.fr"))

    for label, value in pairs:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)
        row += 1

    crit = _entities_summary_lines(entities)
    if crit:
        row += 1
        ws.cell(row=row, column=1, value="Critères structurés (rappel)").font = Font(bold=True, size=11)
        row += 1
        for lab, val in crit:
            ws.cell(row=row, column=1, value=lab).font = bold
            ws.cell(row=row, column=2, value=val)
            row += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 72


def _fill_synthese_sheet(ws, df_raw: pd.DataFrame, intent: str, entities: dict[str, Any] | None) -> None:
    row = 1
    for title, block_df in _build_synthese_dataframes(df_raw, intent, entities):
        row = _append_title(ws, row, title)
        row = _append_dataframe_block(ws, row, block_df)
        row += 1
    _autosize_columns(ws)


def generate_excel(
    results: SearchResults,
    *,
    query_text: str = "",
    intent: str = "recherche_entreprise",
    entities: dict[str, Any] | None = None,
    credits_used: int | None = None,
) -> str:
    """Génère un classeur Excel (résultats + info + synthèse) et renvoie le chemin."""
    df_display = _results_to_dataframe(results, rename=True)
    df_raw = _results_to_dataframe(results, rename=False)

    filename = f"monv_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}.xlsx"
    filepath = Path(settings.EXPORTS_DIR) / filename

    with pd.ExcelWriter(str(filepath), engine="openpyxl") as writer:
        df_display.to_excel(writer, sheet_name="Résultats", index=False)
        ws = writer.sheets["Résultats"]
        _style_header_row(ws, 1)
        for col_idx, col in enumerate(df_display.columns, 1):
            max_len = max(
                len(str(col)),
                int(df_display[col].astype(str).str.len().max()) if len(df_display) > 0 else 0,
            )
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max_len + 4, 44)
        ws.freeze_panes = "A2"

        wb = writer.book
        wb.create_sheet("Info", 0)
        wb.create_sheet("Synthèse", 1)
        _fill_info_sheet(
            wb["Info"],
            query_text=query_text,
            intent=intent,
            entities=entities,
            total=results.total,
            credits_used=credits_used,
        )
        _fill_synthese_sheet(wb["Synthèse"], df_raw, intent, entities)

    return str(filepath)


def generate_csv(results: SearchResults) -> str:
    """Génère un CSV et renvoie le chemin (mêmes colonnes que l'onglet Résultats)."""
    df = _results_to_dataframe(results, rename=True)
    filename = f"monv_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}.csv"
    filepath = Path(settings.EXPORTS_DIR) / filename
    df.to_csv(str(filepath), index=False, encoding="utf-8-sig")
    return str(filepath)
